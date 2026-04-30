from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .base import json_payload, stable_record_hash


TABLE_NAME = "source_comtrade_flows"
CONFLICT_COLUMNS = ("tenant_id", "source_record_hash")


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_int(value: object) -> int | None:
    raw = _clean_text(value)
    if not raw:
        return None
    try:
        return int(float(raw))
    except ValueError:
        return None


def _parse_float(value: object) -> float | None:
    raw = _clean_text(value)
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def _parse_bool(value: object) -> bool | None:
    if isinstance(value, bool):
        return value
    raw = _clean_text(value).lower()
    if not raw:
        return None
    if raw in {"true", "1", "yes", "y"}:
        return True
    if raw in {"false", "0", "no", "n"}:
        return False
    return None


def _code_prefix(value: object, length: int) -> str:
    code = _clean_text(value)
    return code[: min(len(code), length)]


def transform_comtrade_row(tenant_id: str, source_file_name: str, row: dict) -> dict:
    cmd_code = _clean_text(row.get("cmdCode"))
    record_hash = stable_record_hash(
        tenant_id,
        row.get("refYear"),
        row.get("flowCode"),
        row.get("reporterISO"),
        row.get("partnerISO"),
        row.get("cmdCode"),
    )
    return {
        "tenant_id": tenant_id,
        "source_record_hash": record_hash,
        "type_code": _clean_text(row.get("typeCode")),
        "freq_code": _clean_text(row.get("freqCode")),
        "ref_period_id": _parse_int(row.get("refPeriodId")),
        "ref_year": int(row["refYear"]),
        "ref_month": _parse_int(row.get("refMonth")),
        "period": _clean_text(row.get("period")),
        "reporter_code": _parse_int(row.get("reporterCode")),
        "reporter_iso": _clean_text(row.get("reporterISO")),
        "reporter_desc": _clean_text(row.get("reporterDesc")),
        "flow_code": _clean_text(row.get("flowCode")),
        "flow_desc": _clean_text(row.get("flowDesc")),
        "partner_code": _parse_int(row.get("partnerCode")),
        "partner_iso": _clean_text(row.get("partnerISO")),
        "partner_desc": _clean_text(row.get("partnerDesc")),
        "partner2_code": _parse_int(row.get("partner2Code")),
        "partner2_iso": _clean_text(row.get("partner2ISO")),
        "partner2_desc": _clean_text(row.get("partner2Desc")),
        "classification_code": _clean_text(row.get("classificationCode")),
        "classification_search_code": _clean_text(row.get("classificationSearchCode")),
        "is_original_classification": _parse_bool(row.get("isOriginalClassification")),
        "cmd_code": cmd_code,
        "cmd_code_level2": _code_prefix(cmd_code, 2),
        "cmd_code_level4": _code_prefix(cmd_code, 4),
        "cmd_code_level6": _code_prefix(cmd_code, 6),
        "cmd_desc": _clean_text(row.get("cmdDesc")),
        "aggr_level": _parse_int(row.get("aggrLevel")),
        "is_leaf": _parse_bool(row.get("isLeaf")),
        "customs_code": _clean_text(row.get("customsCode")),
        "customs_desc": _clean_text(row.get("customsDesc")),
        "mos_code": _clean_text(row.get("mosCode")),
        "mot_code": _parse_int(row.get("motCode")),
        "mot_desc": _clean_text(row.get("motDesc")),
        "qty_unit_code": _parse_int(row.get("qtyUnitCode")),
        "qty_unit_abbr": _clean_text(row.get("qtyUnitAbbr")),
        "qty": _parse_float(row.get("qty")),
        "is_qty_estimated": _parse_bool(row.get("isQtyEstimated")),
        "alt_qty_unit_code": _parse_int(row.get("altQtyUnitCode")),
        "alt_qty_unit_abbr": _clean_text(row.get("altQtyUnitAbbr")),
        "alt_qty": _parse_float(row.get("altQty")),
        "is_alt_qty_estimated": _parse_bool(row.get("isAltQtyEstimated")),
        "net_wgt": _parse_float(row.get("netWgt")),
        "is_net_wgt_estimated": _parse_bool(row.get("isNetWgtEstimated")),
        "gross_wgt": _parse_float(row.get("grossWgt")),
        "is_gross_wgt_estimated": _parse_bool(row.get("isGrossWgtEstimated")),
        "cifvalue": _parse_float(row.get("cifvalue")),
        "fobvalue": _parse_float(row.get("fobvalue")),
        "primary_value": _parse_float(row.get("primaryValue")),
        "legacy_estimation_flag": _parse_int(row.get("legacyEstimationFlag")),
        "is_reported": _parse_bool(row.get("isReported")),
        "is_aggregate": _parse_bool(row.get("isAggregate")),
        "source_file_name": source_file_name,
        "raw_payload": json_payload(row),
    }


def iter_comtrade_records(tenant_id: str, source_path: Path):
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Sheet1"]
        headers = [str(cell) if cell is not None else "" for cell in next(sheet.iter_rows(values_only=True))]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values, strict=False))
            if row.get("refYear") is None:
                continue
            yield transform_comtrade_row(tenant_id, source_path.name, row)
    finally:
        workbook.close()


