from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from .base import json_payload, stable_record_hash


TABLE_NAME = "source_fda_warning_letters"
CONFLICT_COLUMNS = ("tenant_id", "source_record_hash")

_NON_ALNUM_RE = re.compile(r"[^a-z0-9]+")


def _normalize_text(value: object) -> str:
    return " ".join(_NON_ALNUM_RE.sub(" ", str(value or "").strip().lower()).split())


def _parse_us_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, "%m/%d/%Y").date()
    except ValueError:
        return None


def classify_warning_letter(subject: str, issuing_office: str = "") -> tuple[str, str]:
    lowered = f"{subject} {issuing_office}".lower()
    if any(token in lowered for token in ("adulterated", "sterility", "contamination", "aseptic", "cgm", "cgmp")):
        return "manufacturing_quality", "high"
    if any(token in lowered for token in ("device", "medical device", "design control")):
        return "device_quality", "medium"
    if any(token in lowered for token in ("label", "labeling", "misbranding")):
        return "labeling_compliance", "medium"
    if any(token in lowered for token in ("food", "seafood", "import")):
        return "import_food_safety", "medium"
    return "regulatory_compliance", "low"


def transform_fda_row(tenant_id: str, source_file_name: str, row: dict) -> dict:
    risk_category, severity = classify_warning_letter(
        str(row.get("Subject", "") or ""),
        str(row.get("Issuing Office", "") or ""),
    )
    record_hash = stable_record_hash(
        tenant_id,
        row.get("Company Name"),
        row.get("Posted Date"),
        row.get("Subject"),
    )
    return {
        "tenant_id": tenant_id,
        "source_record_hash": record_hash,
        "posted_date": _parse_us_date(row.get("Posted Date")),
        "letter_issue_date": _parse_us_date(row.get("Letter Issue Date")),
        "company_name": str(row.get("Company Name", "") or "").strip(),
        "company_name_normalized": _normalize_text(row.get("Company Name")),
        "issuing_office": str(row.get("Issuing Office", "") or "").strip(),
        "subject": str(row.get("Subject", "") or "").strip(),
        "risk_category": risk_category,
        "severity": severity,
        "response_letter": str(row.get("Response Letter", "") or "").strip(),
        "closeout_letter": str(row.get("Closeout Letter", "") or "").strip(),
        "source_file_name": source_file_name,
        "raw_payload": json_payload(row),
    }


def iter_fda_records(tenant_id: str, source_path: Path):
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Warning Letter Solr Index"]
        headers = [str(cell) if cell is not None else "" for cell in next(sheet.iter_rows(values_only=True))]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values, strict=False))
            if not str(row.get("Company Name", "") or "").strip():
                continue
            yield transform_fda_row(tenant_id, source_path.name, row)
    finally:
        workbook.close()


