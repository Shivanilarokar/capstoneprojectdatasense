from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from .base import json_payload


TABLE_NAME = "source_ofac_sdn_entities"
CONFLICT_COLUMNS = ("tenant_id", "source_entity_id")

_VALUE_SPLIT_RE = re.compile(r"[;\n|]+")


def _split_multi_value(raw: object) -> list[str]:
    values: list[str] = []
    for part in _VALUE_SPLIT_RE.split(str(raw or "")):
        cleaned = part.strip()
        cleaned = re.sub(r"(?i)^(?:a\.?\s*k\.?\s*a\.?|f\.?\s*k\.?\s*a\.?|n\.?\s*k\.?\s*a\.?)\s*:\s*", "", cleaned)
        if cleaned and cleaned not in values:
            values.append(cleaned)
    return values


def _parse_iso_date(raw: object) -> date | None:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw

    value = str(raw or "").strip()
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def transform_ofac_bis_row(tenant_id: str, source_file_name: str, row: dict) -> dict:
    aliases = _split_multi_value(row.get("Aliases"))
    addresses = _split_multi_value(row.get("Address(es)"))
    return {
        "tenant_id": tenant_id,
        "source_entity_id": str(row.get("Entity ID", "") or "").strip(),
        "primary_name": str(row.get("Primary Name", "") or "").strip(),
        "entity_type": str(row.get("Entity Type", "") or "").strip(),
        "source_list_name": "ofac_sdn",
        "sanctions_programs": str(row.get("Sanctions Program(s)", "") or "").strip(),
        "sanctions_type": str(row.get("Sanctions Type", "") or "").strip(),
        "date_published": _parse_iso_date(row.get("Date Published")),
        "aliases": str(row.get("Aliases", "") or "").strip(),
        "alias_count": len(aliases),
        "date_of_birth": _parse_iso_date(row.get("Date of Birth")),
        "place_of_birth": str(row.get("Place of Birth", "") or "").strip(),
        "nationality": str(row.get("Nationality", "") or "").strip(),
        "citizenship": str(row.get("Citizenship", "") or "").strip(),
        "gender": str(row.get("Gender", "") or "").strip(),
        "address_text": str(row.get("Address(es)", "") or "").strip(),
        "address_count": len(addresses),
        "document_ids": str(row.get("Document IDs", "") or "").strip(),
        "source_file_name": source_file_name,
        "raw_payload": json_payload(row),
    }


def iter_ofac_bis_records(tenant_id: str, source_path: Path):
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet = workbook["SDN Entities"]
        headers = [str(cell) if cell is not None else "" for cell in next(sheet.iter_rows(values_only=True))]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values, strict=False))
            if not str(row.get("Entity ID", "") or "").strip():
                continue
            yield transform_ofac_bis_row(tenant_id, source_path.name, row)
    finally:
        workbook.close()


