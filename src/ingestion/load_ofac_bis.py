from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .base import json_payload


TABLE_NAME = "source_ofac_bis_entities"
CONFLICT_COLUMNS = ("tenant_id", "source_entity_id")


def transform_ofac_bis_row(tenant_id: str, source_file_name: str, row: dict) -> dict:
    return {
        "tenant_id": tenant_id,
        "source_entity_id": str(row.get("Entity ID", "") or "").strip(),
        "primary_name": str(row.get("Primary Name", "") or "").strip(),
        "entity_type": str(row.get("Entity Type", "") or "").strip(),
        "sanctions_programs": str(row.get("Sanctions Program(s)", "") or "").strip(),
        "sanctions_type": str(row.get("Sanctions Type", "") or "").strip(),
        "date_published": str(row.get("Date Published", "") or "").strip(),
        "aliases": str(row.get("Aliases", "") or "").strip(),
        "nationality": str(row.get("Nationality", "") or "").strip(),
        "citizenship": str(row.get("Citizenship", "") or "").strip(),
        "address_text": str(row.get("Address(es)", "") or "").strip(),
        "document_ids": str(row.get("Document IDs", "") or "").strip(),
        "source_file_name": source_file_name,
        "raw_payload": json_payload(row),
    }


def iter_ofac_bis_records(tenant_id: str, source_path: Path):
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet = workbook["SDN Entities"]
        headers = [str(cell) if cell is not None else "" for cell in next(sheet.iter_rows(values_only=True))]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values, strict=False))
            if not str(row.get("Entity ID", "") or "").strip():
                continue
            yield transform_ofac_bis_row(tenant_id, source_path.name, row)
    finally:
        workbook.close()
