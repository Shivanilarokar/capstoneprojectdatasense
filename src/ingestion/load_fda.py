from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .base import json_payload, stable_record_hash


TABLE_NAME = "source_fda_warning_letters"
CONFLICT_COLUMNS = ("tenant_id", "source_record_hash")


def transform_fda_row(tenant_id: str, source_file_name: str, row: dict) -> dict:
    record_hash = stable_record_hash(
        tenant_id,
        row.get("Company Name"),
        row.get("Posted Date"),
        row.get("Subject"),
    )
    return {
        "tenant_id": tenant_id,
        "source_record_hash": record_hash,
        "posted_date": str(row.get("Posted Date", "") or "").strip(),
        "letter_issue_date": str(row.get("Letter Issue Date", "") or "").strip(),
        "company_name": str(row.get("Company Name", "") or "").strip(),
        "issuing_office": str(row.get("Issuing Office", "") or "").strip(),
        "subject": str(row.get("Subject", "") or "").strip(),
        "response_letter": str(row.get("Response Letter", "") or "").strip(),
        "closeout_letter": str(row.get("Closeout Letter", "") or "").strip(),
        "source_file_name": source_file_name,
        "raw_payload": json_payload(row),
    }


def iter_fda_records(tenant_id: str, source_path: Path):
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Warning Letter Solr Index"]
        headers = [str(cell) if cell is not None else "" for cell in next(sheet.iter_rows(values_only=True))]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values, strict=False))
            if not str(row.get("Company Name", "") or "").strip():
                continue
            yield transform_fda_row(tenant_id, source_path.name, row)
    finally:
        workbook.close()
