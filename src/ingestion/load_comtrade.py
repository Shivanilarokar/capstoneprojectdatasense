from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .base import json_payload, stable_record_hash


TABLE_NAME = "source_comtrade_flows"
CONFLICT_COLUMNS = ("tenant_id", "source_record_hash")


def transform_comtrade_row(tenant_id: str, source_file_name: str, row: dict) -> dict:
    record_hash = stable_record_hash(
        tenant_id,
        row.get("refYear"),
        row.get("flowCode"),
        row.get("reporterISO"),
        row.get("partnerISO"),
        row.get("cmdCode"),
    )
    return {
        "tenant_id": tenant_id,
        "source_record_hash": record_hash,
        "ref_year": int(row["refYear"]),
        "flow_code": str(row.get("flowCode", "") or "").strip(),
        "flow_desc": str(row.get("flowDesc", "") or "").strip(),
        "reporter_iso": str(row.get("reporterISO", "") or "").strip(),
        "reporter_desc": str(row.get("reporterDesc", "") or "").strip(),
        "partner_iso": str(row.get("partnerISO", "") or "").strip(),
        "partner_desc": str(row.get("partnerDesc", "") or "").strip(),
        "cmd_code": str(row.get("cmdCode", "") or "").strip(),
        "cmd_desc": str(row.get("cmdDesc", "") or "").strip(),
        "qty": float(row["qty"]) if row.get("qty") is not None else None,
        "net_wgt": float(row["netWgt"]) if row.get("netWgt") is not None else None,
        "cifvalue": float(row["cifvalue"]) if row.get("cifvalue") is not None else None,
        "fobvalue": float(row["fobvalue"]) if row.get("fobvalue") is not None else None,
        "primary_value": float(row["primaryValue"]) if row.get("primaryValue") is not None else None,
        "source_file_name": source_file_name,
        "raw_payload": json_payload(row),
    }


def iter_comtrade_records(tenant_id: str, source_path: Path):
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Sheet1"]
        headers = [str(cell) if cell is not None else "" for cell in next(sheet.iter_rows(values_only=True))]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values, strict=False))
            if row.get("refYear") is None:
                continue
            yield transform_comtrade_row(tenant_id, source_path.name, row)
    finally:
        workbook.close()
